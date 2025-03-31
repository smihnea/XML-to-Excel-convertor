import os
import sys
import glob
import zipfile
import io
import re
import json
import datetime
import time
import traceback
import logging
import shutil
from pathlib import Path
from lxml import etree
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from tkinter import StringVar, Menu
from threading import Thread
from tkcalendar import DateEntry  # You might need to install this: pip install tkcalendar

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("invoice_processor.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("InvoiceProcessor")

# Constants
APP_TITLE = "Invoice Dashboard"
DARK_BG = "#1E1E1E"
DARKER_BG = "#252526"
ITEM_BG = "#2D2D30"
ITEM_BG_ALT = "#252526"
TEXT_COLOR = "#CCCCCC"
ACCENT_COLOR = "#3C7EBF"
ACCENT_HOVER = "#5294D2"
BORDER_COLOR = "#3F3F3F"

# Remove or modify status colors to be much more subtle or transparent
STATUS_COLORS = {
    "Paid": ITEM_BG,       # Same as normal background
    "Pending": ITEM_BG,    # Same as normal background
    "Overdue": ITEM_BG,    # Same as normal background
}

class InvoiceData:
    """Class to manage invoice data and operations"""
    def __init__(self):
        self.invoices = []
        self.filtered_invoices = []
        # Remove 'Folder Name' from columns list
        self.columns = [
            "Nr. doc.", "Data emiterii", "Termen plata",
            "Cota TVA", "Furnizor", "CIF", "Reg. com.", "Adresa", "Judet",
            "IBAN", "Banca", "Produse/Servicii", "Descriere", "U.M.", "Cant.",
            "Pret fara TVA (RON)", "Valoare", "Valoare TVA", "Total", "Total factura"
        ]
        self.ns = {
            "cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2",
            "cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2"
        }
        self.data_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "invoice_data.json")
        self.load_cached_data()
        
        # Create an invoice lookup dictionary for faster existence checks
        self.invoice_lookup = self._build_invoice_lookup()
    
    def _build_invoice_lookup(self):
        """Build a lookup dictionary for faster invoice existence checks"""
        lookup = {}
        for inv in self.invoices:
            doc_num = inv.get("Nr. doc.", "")
            prod_serv = inv.get("Produse/Servicii", "")
            quantity = inv.get("Cant.", "")
            price = inv.get("Pret fara TVA (RON)", "")
            # Create a more unique key that includes quantity and price
            key = f"{doc_num}_{prod_serv}_{quantity}_{price}"
            lookup[key] = True
        return lookup
    
    def load_cached_data(self):
        """Load previously processed invoice data from JSON file if it exists"""
        try:
            if os.path.exists(self.data_file):
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    self.invoices = json.load(f)
                logger.info(f"Loaded {len(self.invoices)} invoices from cache")
                self.filtered_invoices = self.invoices.copy()
            else:
                logger.info("No cache file found, starting with empty invoice list")
        except Exception as e:
            logger.error(f"Error loading cached data: {e}")
            self.invoices = []
            self.filtered_invoices = []
    
    def save_cached_data(self):
        """Save processed invoice data to JSON file"""
        try:
            # Create a backup of the existing file if it exists
            if os.path.exists(self.data_file):
                backup_file = f"{self.data_file}.bak"
                shutil.copy2(self.data_file, backup_file)
                logger.info(f"Created backup of cache file: {backup_file}")
            
            with open(self.data_file, 'w', encoding='utf-8') as f:
                json.dump(self.invoices, f, ensure_ascii=False)
            logger.info(f"Saved {len(self.invoices)} invoices to cache")
            
            # Rebuild the lookup dictionary after saving
            self.invoice_lookup = self._build_invoice_lookup()
        except Exception as e:
            logger.error(f"Error saving cached data: {e}")
            traceback.print_exc()
    
    def get_text(self, element):
        """Safely return element.text or empty string."""
        return element.text if element is not None else ""
    
    def determine_status(self, issue_date, due_date):
        """Determine invoice status based on dates"""
        try:
            today = datetime.datetime.now().date()
            
            if not due_date:
                return "Pending"
                
            due_date_obj = datetime.datetime.strptime(due_date, "%Y-%m-%d").date()
            
            if due_date_obj < today:
                return "Overdue"
            else:
                # Randomly assign Paid or Pending for demonstration
                import random
                return random.choice(["Paid", "Pending"])
        except Exception as e:
            logger.warning(f"Error determining status: {e}. Using 'Pending' as default.")
            return "Pending"
            
    def extract_vat_percent(self, root_xml, invoice_number, filename):
        """Extract VAT percentage from XML and ensure it's properly formatted"""
        # First try to get VAT from the main invoice tax subtotal section
        try:
            # Look for the standard tax rate (S) first
            tax_subtotals = root_xml.findall(".//cac:TaxSubtotal", self.ns)
            for subtotal in tax_subtotals:
                category = subtotal.find("./cac:TaxCategory", self.ns)
                if category is not None:
                    category_id = self.get_text(category.find("./cbc:ID", self.ns))
                    if category_id == "S":  # Standard tax rate
                        vat_percent_text = self.get_text(category.find("./cbc:Percent", self.ns))
                        if vat_percent_text:
                            return self.format_vat_percent(vat_percent_text, filename)
            
            # If no standard rate found, take the first available
            for subtotal in tax_subtotals:
                vat_percent_text = self.get_text(subtotal.find(".//cbc:Percent", self.ns))
                if vat_percent_text:
                    return self.format_vat_percent(vat_percent_text, filename)
            
            # If still not found, try invoice lines
            invoice_lines = root_xml.findall(".//cac:InvoiceLine", self.ns)
            for line in invoice_lines:
                tax_category = line.find(".//cac:ClassifiedTaxCategory", self.ns)
                if tax_category is not None:
                    category_id = self.get_text(tax_category.find("./cbc:ID", self.ns))
                    if category_id == "S":  # Standard tax rate
                        vat_percent_text = self.get_text(tax_category.find("./cbc:Percent", self.ns))
                        if vat_percent_text:
                            return self.format_vat_percent(vat_percent_text, filename)
            
            # If still not found, take the first available from invoice lines
            for line in invoice_lines:
                vat_percent_text = self.get_text(line.find(".//cac:ClassifiedTaxCategory/cbc:Percent", self.ns))
                if vat_percent_text:
                    return self.format_vat_percent(vat_percent_text, filename)
            
            # If no VAT rate found, use default
            logger.warning(f"No VAT rate found for invoice {invoice_number} in {filename}, using default")
            return "0.00"
        except Exception as e:
            logger.error(f"Error extracting VAT rate for invoice {invoice_number} in {filename}: {e}")
            return "0.00"
        
    def format_vat_percent(self, vat_percent_text, filename):
        """Format VAT percentage as a string with 2 decimal places"""
        try:
            # Try to convert to float and format with 2 decimal places
            vat_percent = float(vat_percent_text.replace(',', '.'))
            return f"{vat_percent:.2f}"  # Format as string with 2 decimal places
        except (ValueError, TypeError) as e:
            logger.warning(f"Invalid VAT rate in {filename}: {vat_percent_text}, using default")
            return "0.00"
    
    def process_xml_file(self, xml_content, folder_name, filename):
        """Process a single XML file and extract invoice data"""
        try:
            parser = etree.XMLParser(ns_clean=True, recover=True, encoding='utf-8')
            
            # Try to parse the XML content
            try:
                tree = etree.parse(io.BytesIO(xml_content), parser)
                root_xml = tree.getroot()
            except Exception as e:
                logger.error(f"Error parsing XML file {filename}: {e}")
                return []
            
            # Extract basic invoice data
            invoice_number = self.get_text(root_xml.find(".//cbc:ID", self.ns))
            invoice_number = re.sub(r'\s*nr\.\s*', '', invoice_number, flags=re.IGNORECASE)
            
            issue_date = self.get_text(root_xml.find(".//cbc:IssueDate", self.ns))
            due_date = self.get_text(root_xml.find(".//cbc:DueDate", self.ns))
            
            # Extract VAT percentage using the new function
            vat_percent_formatted = self.extract_vat_percent(root_xml, invoice_number, filename)
            
            # Supplier info
            supplier_name = self.get_text(root_xml.find(
                ".//cac:AccountingSupplierParty//cac:PartyLegalEntity/cbc:RegistrationName", self.ns))
            supplier_cif = self.get_text(root_xml.find(
                ".//cac:AccountingSupplierParty//cac:PartyTaxScheme/cbc:CompanyID", self.ns))
            supplier_regcom = self.get_text(root_xml.find(
                ".//cac:AccountingSupplierParty//cac:PartyLegalEntity/cbc:CompanyID", self.ns))
            supplier_street = self.get_text(root_xml.find(
                ".//cac:AccountingSupplierParty//cac:PostalAddress/cbc:StreetName", self.ns))
            supplier_judet = self.get_text(root_xml.find(
                ".//cac:AccountingSupplierParty//cac:PostalAddress/cbc:CountrySubentity", self.ns))
            
            # Payment details
            payee_account = root_xml.find(".//cac:PaymentMeans/cac:PayeeFinancialAccount", self.ns)
            iban = self.get_text(payee_account.find("cbc:ID", self.ns)) if payee_account is not None else ""
            bank_name = self.get_text(payee_account.find("cbc:Name", self.ns)) if payee_account is not None else ""
            
            # Total invoice
            total_invoice = self.get_text(root_xml.find(".//cac:LegalMonetaryTotal/cbc:PayableAmount", self.ns))
            
            # Process invoice lines
            invoice_lines = root_xml.findall(".//cac:InvoiceLine", self.ns)
            
            # If no lines, create a single invoice record
            if not invoice_lines:
                invoice = {
                    "Nr. doc.": invoice_number,
                    "Data emiterii": issue_date,
                    "Termen plata": due_date,
                    "Cota TVA": vat_percent_formatted,  # Use formatted value
                    "Furnizor": supplier_name,
                    "CIF": supplier_cif,
                    "Reg. com.": supplier_regcom,
                    "Adresa": supplier_street,
                    "Judet": supplier_judet,
                    "IBAN": iban,
                    "Banca": bank_name,
                    "Produse/Servicii": "",
                    "Descriere": "",
                    "U.M.": "",
                    "Cant.": "",
                    "Pret fara TVA (RON)": "",
                    "Valoare": "",
                    "Valoare TVA": "",
                    "Total": "",
                    "Total factura": total_invoice,
                    "Status": self.determine_status(issue_date, due_date),
                    "Filename": filename,
                    "Folder Name": folder_name  # Keep for internal reference but not in column list
                }
                return [invoice]
            
            # Otherwise, create one record per line
            invoices = []
            for line in invoice_lines:
                # Quantity & unit
                invoiced_qty = line.find("./cbc:InvoicedQuantity", self.ns)
                quantity_text = self.get_text(invoiced_qty)
                uom = invoiced_qty.get("unitCode") if invoiced_qty is not None else ""
                
                # Unit price
                price_el = line.find("./cac:Price/cbc:PriceAmount", self.ns)
                price_text = self.get_text(price_el)
                
                # Net line extension
                line_ext_el = line.find("./cbc:LineExtensionAmount", self.ns)
                line_ext_text = self.get_text(line_ext_el)
                
                # Get line-specific VAT rate if available, otherwise use the invoice VAT
                line_vat_percent = vat_percent_formatted
                line_tax_category = line.find(".//cac:ClassifiedTaxCategory", self.ns)
                if line_tax_category is not None:
                    line_vat_text = self.get_text(line_tax_category.find("./cbc:Percent", self.ns))
                    if line_vat_text:
                        line_vat_percent = self.format_vat_percent(line_vat_text, filename)
                
                # Calculate line tax & total
                try:
                    line_ext = float(line_ext_text.replace(',', '.')) if line_ext_text else 0.0
                    # Use the numeric value for calculations
                    line_vat_value = float(line_vat_percent)
                    line_tax = round(line_ext * line_vat_value / 100, 2)
                    line_total = round(line_ext + line_tax, 2)
                    
                    line_tax_text = f"{line_tax:.2f}"
                    line_total_text = f"{line_total:.2f}"
                except (ValueError, TypeError) as e:
                    logger.warning(f"Error calculating tax & total: {e}")
                    line_tax_text = "0.00"
                    line_total_text = "0.00"
                
                # Product/Service name
                item_name = self.get_text(line.find("./cac:Item/cbc:Name", self.ns))
                # Item description
                item_desc = self.get_text(line.find("./cac:Item/cbc:Description", self.ns))
                
                invoice = {
                    "Nr. doc.": invoice_number,
                    "Data emiterii": issue_date,
                    "Termen plata": due_date,
                    "Cota TVA": line_vat_percent,  # Use line-specific VAT if available
                    "Furnizor": supplier_name,
                    "CIF": supplier_cif,
                    "Reg. com.": supplier_regcom,
                    "Adresa": supplier_street,
                    "Judet": supplier_judet,
                    "IBAN": iban,
                    "Banca": bank_name,
                    "Produse/Servicii": item_name,
                    "Descriere": item_desc,
                    "U.M.": uom,
                    "Cant.": quantity_text,
                    "Pret fara TVA (RON)": price_text,
                    "Valoare": line_ext_text,
                    "Valoare TVA": line_tax_text,
                    "Total": line_total_text,
                    "Total factura": total_invoice,
                    "Status": self.determine_status(issue_date, due_date),
                    "Filename": filename,
                    "Folder Name": folder_name  # Keep for internal reference but not in column list
                }
                invoices.append(invoice)
            
            return invoices
        except Exception as e:
            logger.error(f"Error processing XML file {filename}: {e}")
            traceback.print_exc()
            return []
    
    def process_folder(self, folder_path, callback=None):
        """Process all XML files in a folder and its subfolders"""
        xml_files = []
        
        # Check if it's a directory or a ZIP file
        if os.path.isdir(folder_path):
            logger.info(f"Processing directory: {folder_path}")
            for root_dir, dirs, files in os.walk(folder_path):
                folder_name = os.path.basename(root_dir)
                for filename in files:
                    if not filename.lower().endswith(".xml"):
                        continue
                    if "semnatura" in filename.lower():
                        continue
                    file_path = os.path.join(root_dir, filename)
                    with open(file_path, "rb") as f:
                        xml_content = f.read()
                    xml_files.append((xml_content, folder_name, filename))
        elif folder_path.lower().endswith(".zip"):
            logger.info(f"Processing ZIP file: {folder_path}")
            xml_files = self.collect_xml_files_from_zip(folder_path)
        
        logger.info(f"Found {len(xml_files)} XML files to process")
        
        # Process each XML file
        new_invoices = []
        for idx, (xml_content, folder_name, filename) in enumerate(xml_files):
            invoices = self.process_xml_file(xml_content, folder_name, filename)
            if invoices:
                # Check if these invoices already exist
                for invoice in invoices:
                    # Use the lookup dictionary for faster existence checking
                    doc_num = invoice.get("Nr. doc.", "")
                    prod_serv = invoice.get("Produse/Servicii", "")
                    quantity = invoice.get("Cant.", "")
                    price = invoice.get("Pret fara TVA (RON)", "")
                    # Use the same key format as in _build_invoice_lookup
                    key = f"{doc_num}_{prod_serv}_{quantity}_{price}"
                    
                    if key not in self.invoice_lookup:
                        new_invoices.append(invoice)
                        self.invoice_lookup[key] = True  # Update lookup dict
            
            # Update progress callback
            if callback:
                progress = (idx + 1) / len(xml_files) * 100
                callback(progress, f"Processing file {idx + 1} of {len(xml_files)}")
        
        # Add new invoices to the list
        if new_invoices:
            logger.info(f"Adding {len(new_invoices)} new invoices")
            self.invoices.extend(new_invoices)
            self.filtered_invoices = self.invoices.copy()
            self.save_cached_data()
        else:
            logger.info("No new invoices found")
        
        return len(new_invoices)
    
    def collect_xml_files_from_zip(self, zip_path):
        """Collect XML files from a ZIP file (including nested ZIPs)"""
        xml_files = []
        
        base = os.path.basename(zip_path)
        current_folder_name = os.path.splitext(base)[0]
        
        try:
            with zipfile.ZipFile(zip_path, 'r') as zf:
                for name in zf.namelist():
                    if name.endswith('/'):
                        continue
                    
                    try:
                        file_bytes = zf.read(name)
                        if name.lower().endswith('.zip'):
                            nested_zip = io.BytesIO(file_bytes)
                            sub_folder_name = f"{current_folder_name}/{os.path.splitext(os.path.basename(name))[0]}"
                            xml_files.extend(self.collect_xml_files_from_inmemory(nested_zip, sub_folder_name))
                        elif name.lower().endswith('.xml') and "semnatura" not in name.lower():
                            filename = os.path.basename(name)
                            xml_files.append((file_bytes, current_folder_name, filename))
                    except Exception as e:
                        logger.error(f"Error reading file {name} from ZIP: {e}")
        except Exception as e:
            logger.error(f"Error opening ZIP file {zip_path}: {e}")
            traceback.print_exc()
        
        logger.info(f"Found {len(xml_files)} XML files in ZIP {zip_path}")
        return xml_files
    
    def collect_xml_files_from_inmemory(self, zip_bytes, current_folder_name):
        """Collect XML files from an in-memory ZIP file"""
        xml_files = []
        
        try:
            with zipfile.ZipFile(zip_bytes, 'r') as zf:
                for name in zf.namelist():
                    if name.endswith('/'):
                        continue
                    
                    try:
                        file_bytes = zf.read(name)
                        if name.lower().endswith('.zip'):
                            nested_zip = io.BytesIO(file_bytes)
                            sub_folder_name = f"{current_folder_name}/{os.path.splitext(os.path.basename(name))[0]}"
                            xml_files.extend(self.collect_xml_files_from_inmemory(nested_zip, sub_folder_name))
                        elif name.lower().endswith('.xml') and "semnatura" not in name.lower():
                            filename = os.path.basename(name)
                            xml_files.append((file_bytes, current_folder_name, filename))
                    except Exception as e:
                        logger.error(f"Error reading file {name} from nested ZIP: {e}")
        except Exception as e:
            logger.error(f"Error opening nested ZIP: {e}")
            traceback.print_exc()
        
        return xml_files
    
    def search_invoices(self, query):
        """Search invoices by any field"""
        if not query:
            self.filtered_invoices = self.invoices.copy()
            return
        
        query = query.lower()
        self.filtered_invoices = []
        
        for invoice in self.invoices:
            for key, value in invoice.items():
                if key in self.columns and value and query in str(value).lower():
                    self.filtered_invoices.append(invoice)
                    break
    
    def filter_invoices(self, status=None, date_from=None, date_to=None):
        """Filter invoices by status and date range"""
        # Start with all invoices
        self.filtered_invoices = self.invoices.copy()
        
        # Filter by status if specified
        if status and status != "All":
            self.filtered_invoices = [inv for inv in self.filtered_invoices if inv.get("Status") == status]
        
        # Filter by date range if specified
        if date_from or date_to:
            filtered = []
            for inv in self.filtered_invoices:
                try:
                    issue_date = inv.get("Data emiterii", "")
                    if not issue_date:
                        continue
                        
                    inv_date = datetime.datetime.strptime(issue_date, "%Y-%m-%d").date()
                    
                    if date_from and date_to:
                        if date_from <= inv_date <= date_to:
                            filtered.append(inv)
                    elif date_from:
                        if date_from <= inv_date:
                            filtered.append(inv)
                    elif date_to:
                        if inv_date <= date_to:
                            filtered.append(inv)
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid date format in invoice: {e}")
                    continue
            
            self.filtered_invoices = filtered
    
    def export_to_excel(self, output_path):
        """Export filtered invoices to Excel"""
        if not self.filtered_invoices:
            logger.warning("No invoices to export")
            return False
        
        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            
            # Add headers
            headers = self.columns.copy()
            ws.append(headers)
            
            # Add data
            for invoice in self.filtered_invoices:
                row = [invoice.get(col, "") for col in headers]
                ws.append(row)
            
            # Adjust column widths
            for col_idx, column in enumerate(ws.columns, 1):
                max_length = 0
                column_name = column[0].value
                
                for i, cell in enumerate(column[:100]):
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = max(max_length, len(str(column_name))) + 3
                ws.column_dimensions[chr(64 + col_idx)].width = min(adjusted_width, 35)
            
            # Save the workbook
            wb.save(output_path)
            logger.info(f"Successfully exported {len(self.filtered_invoices)} invoices to {output_path}")
            return True
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            traceback.print_exc()
            return False
    
    def advanced_filter(self, filters=None):
        """Apply advanced filters to invoices"""
        if not filters or not any(filters.values()):
            self.filtered_invoices = self.invoices.copy()
            return
        
        # Start with all invoices
        self.filtered_invoices = []
        
        for invoice in self.invoices:
            match = True
            
            # Check status filter
            if filters.get('status') and filters['status'] != "All":
                if invoice.get("Status") != filters['status']:
                    match = False
            
            # Check date range filter
            if match and (filters.get('date_from') or filters.get('date_to')):
                try:
                    issue_date = invoice.get("Data emiterii", "")
                    if not issue_date:
                        match = False
                    else:
                        inv_date = datetime.datetime.strptime(issue_date, "%Y-%m-%d").date()
                        
                        if filters.get('date_from') and inv_date < filters['date_from']:
                            match = False
                        
                        if filters.get('date_to') and inv_date > filters['date_to']:
                            match = False
                except (ValueError, TypeError):
                    match = False
            
            # Check text filters
            text_filters = {
                'nr_doc': "Nr. doc.",
                'furnizor': "Furnizor",
                'cif': "CIF",
                'adresa': "Adresa"
            }
            
            for filter_key, invoice_key in text_filters.items():
                if match and filters.get(filter_key):
                    value = str(invoice.get(invoice_key, "")).lower()
                    if filters[filter_key].lower() not in value:
                        match = False
            
            # Check amount filters
            if match and filters.get('min_amount'):
                try:
                    total = float(invoice.get("Total factura", "0").replace(',', '.'))
                    if total < float(filters['min_amount']):
                        match = False
                except (ValueError, TypeError):
                    pass
                    
            if match and filters.get('max_amount'):
                try:
                    total = float(invoice.get("Total factura", "0").replace(',', '.'))
                    if total > float(filters['max_amount']):
                        match = False
                except (ValueError, TypeError):
                    pass
            
            # If all filters passed, add to filtered list
            if match:
                self.filtered_invoices.append(invoice)


class DarkTheme:
    """Dark theme styling for tkinter widgets"""
    @staticmethod
    def configure_ttk_styles():
        style = ttk.Style()
        
        # Configure TFrame
        style.configure("TFrame", background=DARK_BG)
        
        # Configure TLabel
        style.configure("TLabel", background=DARK_BG, foreground=TEXT_COLOR)
        
        # Configure TButton
        style.configure("TButton", background=ACCENT_COLOR, foreground="white", 
                       borderwidth=0, focusthickness=0, padding=(10, 5))
        style.map("TButton", 
                 background=[("active", ACCENT_HOVER), ("disabled", "#555555")],
                 foreground=[("disabled", "#AAAAAA")])
        
        # Configure Accent.TButton
        style.configure("Accent.TButton", background=ACCENT_COLOR, foreground="white")
        style.map("Accent.TButton", 
                 background=[("active", ACCENT_HOVER), ("disabled", "#555555")],
                 foreground=[("disabled", "#AAAAAA")])
        
        # Configure TEntry
        style.configure("TEntry", fieldbackground=ITEM_BG, foreground=TEXT_COLOR, 
                       borderwidth=1, bordercolor=BORDER_COLOR)
        style.map("TEntry", fieldbackground=[("disabled", "#333333")])
        
        # Configure TCombobox - Fixed the syntax error here
        style.configure("TCombobox", fieldbackground=ITEM_BG, foreground=TEXT_COLOR, 
                       background=DARK_BG, arrowcolor=TEXT_COLOR)
        style.map("TCombobox", fieldbackground=[("readonly", ITEM_BG)], 
                 selectbackground=[("readonly", ACCENT_COLOR)])
        
        # Configure Treeview - IMPORTANT: Modify this to have consistent background
        style.configure("Treeview", 
                       background=ITEM_BG, 
                       foreground=TEXT_COLOR, 
                       fieldbackground=ITEM_BG,
                       borderwidth=0)
        
        # Only change text color on selection, not background
        style.map("Treeview", 
                 background=[("selected", ITEM_BG)],  # Keep background same color
                 foreground=[("selected", "white")])  # Only change text color
        
        # Configure Treeview.Heading
        style.configure("Treeview.Heading", 
                       background=DARKER_BG, 
                       foreground=TEXT_COLOR, 
                       borderwidth=1,
                       relief="flat")
        style.map("Treeview.Heading", 
                 background=[("active", ITEM_BG)])
        
        # Configure Horizontal.TProgressbar
        style.configure("Horizontal.TProgressbar", 
                       background=ACCENT_COLOR, 
                       troughcolor=ITEM_BG, 
                       borderwidth=0)
        
        # Configure TScrollbar
        style.configure("TScrollbar", 
                       background=DARKER_BG, 
                       troughcolor=ITEM_BG, 
                       borderwidth=0,
                       arrowcolor=TEXT_COLOR)
        style.map("TScrollbar", 
                 background=[("active", ACCENT_COLOR), ("disabled", DARKER_BG)])


class InvoiceDashboard(tk.Tk):
    """Main application window"""
    def __init__(self):
        super().__init__()
        
        # Configure window
        self.title(APP_TITLE)
        self.geometry("1200x700")
        self.minsize(800, 600)
        self.configure(bg=DARK_BG)
        
        # Apply dark theme
        DarkTheme.configure_ttk_styles()
        
        # Initialize invoice data
        self.invoice_data = InvoiceData()
        
        # Initialize filter variables - MOVED BEFORE create_ui()
        self.filter_vars = {
            'status': tk.StringVar(value="All"),
            'date_from': None,
            'date_to': None,
            'nr_doc': tk.StringVar(),
            'furnizor': tk.StringVar(),
            'cif': tk.StringVar(),
            'adresa': tk.StringVar(),
            'min_amount': tk.StringVar(),
            'max_amount': tk.StringVar(),
        }
        
        # Create UI - Now filter_vars exists before this is called
        self.create_ui()
        
        # Load initial data
        self.refresh_table()
    
    def create_ui(self):
        """Create the user interface"""
        # Configure grid
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        
        # Create header frame
        self.create_header()
        
        # Create main content frame
        self.content_frame = ttk.Frame(self)
        self.content_frame.grid(row=1, column=0, padx=20, pady=(0, 20), sticky="nsew")
        self.content_frame.grid_columnconfigure(0, weight=1)
        self.content_frame.grid_rowconfigure(1, weight=1)
        
        # Create filter bar
        self.create_filter_bar()
        
        # Create table
        self.create_table()
        
        # Create status bar
        self.create_status_bar()
    
    def create_header(self):
        """Create the header with search and main actions"""
        header_frame = ttk.Frame(self)
        header_frame.grid(row=0, column=0, padx=20, pady=10, sticky="ew")
        header_frame.grid_columnconfigure(0, weight=1)
        
        # Search box with placeholder
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self.on_search)
        
        self.search_entry = ttk.Entry(header_frame, textvariable=self.search_var, width=50)
        self.search_entry.grid(row=0, column=0, padx=(10, 5), pady=10, sticky="ew")
        
        # Set placeholder text
        self.search_placeholder = "Search for invoice..."
        self.search_entry.insert(0, self.search_placeholder)
        self.search_entry.config(foreground="#888888")
        
        # Bind focus events for placeholder handling
        self.search_entry.bind("<FocusIn>", self.on_search_focus_in)
        self.search_entry.bind("<FocusOut>", self.on_search_focus_out)
        
        # Search button
        search_button = ttk.Button(header_frame, text="Search", command=self.on_search_button)
        search_button.grid(row=0, column=1, padx=(5, 10), pady=10)
        
        # Import button
        import_button = ttk.Button(header_frame, text="Import", command=self.on_import)
        import_button.grid(row=0, column=2, padx=5, pady=10)
        
        # Export button
        export_button = ttk.Button(header_frame, text="Export", command=self.on_export)
        export_button.grid(row=0, column=3, padx=5, pady=10)
    
    def on_search_focus_in(self, event):
        """Handle search entry focus in - clear placeholder if needed"""
        if self.search_var.get() == self.search_placeholder:
            self.search_entry.delete(0, tk.END)
            self.search_entry.config(foreground=TEXT_COLOR)
    
    def on_search_focus_out(self, event):
        """Handle search entry focus out - restore placeholder if empty"""
        if not self.search_var.get():
            self.search_entry.insert(0, self.search_placeholder)
            self.search_entry.config(foreground="#888888")
    
    def create_filter_bar(self):
        """Create the filter bar with status and date filters"""
        filter_frame = ttk.Frame(self.content_frame)
        filter_frame.grid(row=0, column=0, padx=0, pady=10, sticky="ew")
        filter_frame.grid_columnconfigure(7, weight=1)
        
        # Status filter
        status_label = ttk.Label(filter_frame, text="Status:")
        status_label.grid(row=0, column=0, padx=(10, 5), pady=5, sticky="w")
        
        status_options = ["All", "Paid", "Pending", "Overdue"]
        status_dropdown = ttk.Combobox(filter_frame, values=status_options, 
                                      textvariable=self.filter_vars['status'], state="readonly", width=15)
        status_dropdown.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        status_dropdown.bind("<<ComboboxSelected>>", self.on_auto_filter)
        
        # Date range filter with calendar widgets
        date_label = ttk.Label(filter_frame, text="Date Range:")
        date_label.grid(row=0, column=2, padx=(20, 5), pady=5, sticky="w")
        
        # Use DateEntry widget for better date selection
        self.date_from_entry = DateEntry(filter_frame, width=15, background=ACCENT_COLOR,
                                         foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd',
                                         selectbackground=ACCENT_HOVER)
        self.date_from_entry.grid(row=0, column=3, padx=5, pady=5, sticky="w")
        self.date_from_entry.delete(0, tk.END)  # Clear the default date
        self.date_from_entry.bind("<<DateEntrySelected>>", self.on_auto_filter)
        
        date_to_label = ttk.Label(filter_frame, text="to")
        date_to_label.grid(row=0, column=4, padx=5, pady=5, sticky="w")
        
        self.date_to_entry = DateEntry(filter_frame, width=15, background=ACCENT_COLOR,
                                       foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd',
                                       selectbackground=ACCENT_HOVER)
        self.date_to_entry.grid(row=0, column=5, padx=5, pady=5, sticky="w")
        self.date_to_entry.delete(0, tk.END)  # Clear the default date
        self.date_to_entry.bind("<<DateEntrySelected>>", self.on_auto_filter)
        
        # Advanced Filters button - moved from header to here
        adv_filter_button = ttk.Button(filter_frame, text="Advanced Filters", command=self.show_advanced_filters)
        adv_filter_button.grid(row=0, column=6, padx=5, pady=5, sticky="w")
        
        # Clear Filters button
        clear_filter_button = ttk.Button(filter_frame, text="Clear Filters", command=self.on_clear_filters)
        clear_filter_button.grid(row=0, column=7, padx=5, pady=5, sticky="w")
        
        # Refresh button
        refresh_button = ttk.Button(filter_frame, text="Refresh", command=self.refresh_table)
        refresh_button.grid(row=0, column=8, padx=5, pady=5, sticky="e")
    
    def create_table(self):
        """Create the invoice table"""
        # Create a frame for the table
        table_frame = ttk.Frame(self.content_frame)
        table_frame.grid(row=1, column=0, sticky="nsew")
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(0, weight=1)
        
        # Add Copy-Paste instructions
        copy_help = ttk.Label(table_frame, text="Right-click or Ctrl+C to copy data", 
                             font=("Segoe UI", 9), foreground="#888888")
        copy_help.grid(row=2, column=0, sticky="w", padx=5)
        
        # Create scrollbars
        vsb = ttk.Scrollbar(table_frame, orient="vertical")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb = ttk.Scrollbar(table_frame, orient="horizontal")
        hsb.grid(row=1, column=0, sticky="ew")
        
        # Create treeview
        self.tree = ttk.Treeview(table_frame, columns=self.invoice_data.columns, 
                                show="headings", selectmode="browse",
                                yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Configure scrollbars
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        
        # Configure column headings
        for col in self.invoice_data.columns:
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_by_column(c))
            width = 100
            if col in ["Furnizor", "Adresa"]:
                width = 150
            elif col in ["Nr. doc.", "Data emiterii", "Termen plata"]:
                width = 100
            elif col in ["Cota TVA", "Cant.", "Valoare TVA"]:
                width = 80
            elif col in ["Pret fara TVA (RON)", "Valoare", "Total", "Total factura"]:
                width = 120
            
            self.tree.column(col, width=width, minwidth=50)
        
        # Place the treeview
        self.tree.grid(row=0, column=0, sticky="nsew")
        
        # Configure tag colors for status - all set to same background to avoid coloring rows
        self.tree.tag_configure("Paid", background=ITEM_BG)
        self.tree.tag_configure("Pending", background=ITEM_BG)
        self.tree.tag_configure("Overdue", background=ITEM_BG)
        
        # Bind double-click event
        self.tree.bind("<Double-1>", self.on_invoice_double_click)
        
        # Enable copy functionality
        self.tree.bind("<Control-c>", self.copy_selection)
        self.tree.bind("<Button-3>", self.show_context_menu)
        
        # Create context menu
        self.context_menu = Menu(self, tearoff=0)
        self.context_menu.add_command(label="Copy", command=self.copy_selection)
        self.context_menu.add_command(label="Copy All Row", command=self.copy_row)
        
        # Create pagination frame
        pagination_frame = ttk.Frame(self.content_frame)
        pagination_frame.grid(row=2, column=0, pady=10, sticky="ew")
        pagination_frame.grid_columnconfigure(5, weight=1)
        
        # Previous page button
        self.prev_button = ttk.Button(pagination_frame, text="Prev", command=self.on_prev_page, width=10)
        self.prev_button.grid(row=0, column=0, padx=5, pady=5)
        
        # Page indicator
        self.page_var = tk.StringVar(value="Page 1")
        page_label = ttk.Label(pagination_frame, textvariable=self.page_var)
        page_label.grid(row=0, column=1, padx=10, pady=5)
        
        # Next page button
        self.next_button = ttk.Button(pagination_frame, text="Next", command=self.on_next_page, width=10)
        self.next_button.grid(row=0, column=2, padx=5, pady=5)
        
        # Items per page
        items_label = ttk.Label(pagination_frame, text="Items per page:")
        items_label.grid(row=0, column=3, padx=(20, 5), pady=5)
        
        self.page_size_var = tk.StringVar(value="25")
        page_size_options = ["10", "25", "50", "100", "All"]
        page_size_dropdown = ttk.Combobox(pagination_frame, values=page_size_options, 
                                         textvariable=self.page_size_var, state="readonly", width=5)
        page_size_dropdown.grid(row=0, column=4, padx=5, pady=5)
        page_size_dropdown.bind("<<ComboboxSelected>>", self.on_page_size_change)
        
        # Showing entries label
        self.entries_var = tk.StringVar(value="Showing 0-0 of 0 entries")
        entries_label = ttk.Label(pagination_frame, textvariable=self.entries_var)
        entries_label.grid(row=0, column=6, padx=10, pady=5, sticky="e")
        
        # Initialize pagination variables
        self.current_page = 1
        self.page_size = 25  # Set default to match dropdown
    
    def create_status_bar(self):
        """Create the status bar"""
        status_frame = ttk.Frame(self)
        status_frame.grid(row=2, column=0, padx=20, pady=5, sticky="ew")
        
        self.status_var = tk.StringVar(value="Displaying 0 invoices")
        status_label = ttk.Label(status_frame, textvariable=self.status_var)
        status_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
    
    def refresh_table(self):
        """Refresh the table with current data"""
        # Clear the table
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Calculate pagination
        total_invoices = len(self.invoice_data.filtered_invoices)
        
        if self.page_size_var.get() == "All":
            self.page_size = total_invoices
            start_idx = 0
            end_idx = total_invoices
        else:
            self.page_size = int(self.page_size_var.get())
            max_pages = max(1, (total_invoices + self.page_size - 1) // self.page_size)
            
            if self.current_page > max_pages:
                self.current_page = max_pages
            
            start_idx = (self.current_page - 1) * self.page_size
            end_idx = min(start_idx + self.page_size, total_invoices)
        
        # Update pagination controls
        self.page_var.set(f"Page {self.current_page}")
        if total_invoices > 0:
            self.entries_var.set(f"Showing {start_idx + 1}-{end_idx} of {total_invoices} entries")
        else:
            self.entries_var.set("Showing 0-0 of 0 entries")
        
        # Enable/disable pagination buttons
        self.prev_button.configure(state="normal" if self.current_page > 1 else "disabled")
        self.next_button.configure(state="normal" if end_idx < total_invoices else "disabled")
        
        # Add data to the table - Always add the status tag, but all status tags now have the same background
        for i, invoice in enumerate(self.invoice_data.filtered_invoices[start_idx:end_idx]):
            values = [invoice.get(col, "") for col in self.invoice_data.columns]
            status = invoice.get("Status", "")
            item_id = self.tree.insert("", "end", values=values, tags=(status,))
        
        # Update status
        self.status_var.set(f"Displaying {end_idx - start_idx} invoices")
        logger.debug(f"Table refreshed, displaying {end_idx - start_idx} invoices")

    def show_advanced_filters(self):
        """Show advanced filter dialog"""
        filter_dialog = tk.Toplevel(self)
        filter_dialog.title("Advanced Filters")
        filter_dialog.geometry("600x550")
        filter_dialog.configure(bg=DARK_BG)
        filter_dialog.transient(self)
        filter_dialog.grab_set()
        
        # Configure grid
        filter_dialog.grid_columnconfigure(0, weight=1)
        
        # Create notebook with tabs
        notebook = ttk.Notebook(filter_dialog)
        notebook.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
        filter_dialog.grid_rowconfigure(0, weight=1)
        
        # Create filters tab
        filters_frame = ttk.Frame(notebook)
        notebook.add(filters_frame, text="Filters")
        
        # Configure grid for filters frame
        filters_frame.grid_columnconfigure(1, weight=1)
        
        # Create form with all filter options
        row = 0
        
        # Status filter
        ttk.Label(filters_frame, text="Status:").grid(row=row, column=0, padx=10, pady=5, sticky="w")
        status_options = ["All", "Paid", "Pending", "Overdue"]
        status_dropdown = ttk.Combobox(filters_frame, values=status_options, 
                                      textvariable=self.filter_vars['status'], state="readonly", width=20)
        status_dropdown.grid(row=row, column=1, padx=10, pady=5, sticky="ew")
        row += 1
        
        # Date range with calendar widgets
        ttk.Label(filters_frame, text="Date From:").grid(row=row, column=0, padx=10, pady=5, sticky="w")
        # Custom date entry implementation that allows empty values
        date_from_frame = ttk.Frame(filters_frame)
        date_from_frame.grid(row=row, column=1, padx=10, pady=5, sticky="ew")
        
        self.adv_date_from = DateEntry(date_from_frame, width=20, background=ACCENT_COLOR,
                             foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd',
                             selectbackground=ACCENT_HOVER)
        self.adv_date_from.pack(side="left", fill="x", expand=True)
        self.adv_date_from.delete(0, tk.END)  # Clear the default date
        
        clear_from_btn = ttk.Button(date_from_frame, text="×", width=3, 
                                    command=lambda: self.adv_date_from.delete(0, tk.END))
        clear_from_btn.pack(side="right", padx=(5, 0))
        row += 1
        
        ttk.Label(filters_frame, text="Date To:").grid(row=row, column=0, padx=10, pady=5, sticky="w")
        # Custom date entry implementation that allows empty values
        date_to_frame = ttk.Frame(filters_frame)
        date_to_frame.grid(row=row, column=1, padx=10, pady=5, sticky="ew")
        
        self.adv_date_to = DateEntry(date_to_frame, width=20, background=ACCENT_COLOR,
                           foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd',
                           selectbackground=ACCENT_HOVER)
        self.adv_date_to.pack(side="left", fill="x", expand=True)
        self.adv_date_to.delete(0, tk.END)  # Clear the default date
        
        clear_to_btn = ttk.Button(date_to_frame, text="×", width=3, 
                                  command=lambda: self.adv_date_to.delete(0, tk.END))
        clear_to_btn.pack(side="right", padx=(5, 0))
        row += 1
        
        # Text filters
        text_filters = [
            ("Invoice Number:", 'nr_doc'),
            ("Supplier:", 'furnizor'),
            ("CIF:", 'cif'),
            ("Address:", 'adresa')
        ]
        
        for label_text, var_name in text_filters:
            ttk.Label(filters_frame, text=label_text).grid(row=row, column=0, padx=10, pady=5, sticky="w")
            ttk.Entry(filters_frame, textvariable=self.filter_vars[var_name], width=30).grid(
                row=row, column=1, padx=10, pady=5, sticky="ew")
            row += 1
        
        # Amount range
        ttk.Label(filters_frame, text="Min Amount:").grid(row=row, column=0, padx=10, pady=5, sticky="w")
        ttk.Entry(filters_frame, textvariable=self.filter_vars['min_amount'], width=20).grid(
            row=row, column=1, padx=10, pady=5, sticky="ew")
        row += 1
        
        ttk.Label(filters_frame, text="Max Amount:").grid(row=row, column=0, padx=10, pady=5, sticky="w")
        ttk.Entry(filters_frame, textvariable=self.filter_vars['max_amount'], width=20).grid(
            row=row, column=1, padx=10, pady=5, sticky="ew")
        row += 1
        
        # Create columns tab
        columns_frame = ttk.Frame(notebook)
        notebook.add(columns_frame, text="Columns")
        
        # Create a canvas with scrollbar for many columns
        canvas = tk.Canvas(columns_frame, bg=DARK_BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(columns_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y", pady=10)
        
        # Column visibility variables and checkboxes
        self.column_vars = {}
        for i, col in enumerate(self.invoice_data.columns):
            var = tk.BooleanVar(value=True)  # All columns are visible by default
            self.column_vars[col] = var
            
            checkbox = ttk.Checkbutton(scrollable_frame, text=col, variable=var)
            checkbox.grid(row=i, column=0, padx=10, pady=5, sticky="w")
        
        # Buttons for both tabs
        button_frame = ttk.Frame(filter_dialog)
        button_frame.grid(row=1, column=0, padx=10, pady=20, sticky="ew")
        
        apply_button = ttk.Button(
            button_frame, 
            text="Apply", 
            command=lambda: self.apply_advanced_filters(
                filter_dialog, 
                self.adv_date_from.get_date() if self.adv_date_from.get() else None,
                self.adv_date_to.get_date() if self.adv_date_to.get() else None
            )
        )
        apply_button.pack(side="right", padx=5)
        
        clear_button = ttk.Button(button_frame, text="Clear All", command=self.clear_advanced_filters)
        clear_button.pack(side="right", padx=5)
        
        cancel_button = ttk.Button(button_frame, text="Cancel", command=filter_dialog.destroy)
        cancel_button.pack(side="right", padx=5)
    
    def apply_advanced_filters(self, dialog, date_from=None, date_to=None):
        """Apply all advanced filters and close the dialog"""
        filters = {
            'status': self.filter_vars['status'].get(),
            'date_from': date_from,
            'date_to': date_to,
            'nr_doc': self.filter_vars['nr_doc'].get(),
            'furnizor': self.filter_vars['furnizor'].get(),
            'cif': self.filter_vars['cif'].get(),
            'adresa': self.filter_vars['adresa'].get(),
            'min_amount': self.filter_vars['min_amount'].get(),
            'max_amount': self.filter_vars['max_amount'].get(),
        }
        
        # Update main filter bar dates
        if date_from:
            self.date_from_entry.set_date(date_from)
        if date_to:
            self.date_to_entry.set_date(date_to)
        
        # Apply filters
        self.invoice_data.advanced_filter(filters)
        
        # Apply column selection
        visible_columns = []
        for col, var in self.column_vars.items():
            if var.get():
                visible_columns.append(col)
        
        # Configure visible columns in the treeview
        if visible_columns:
            for col in self.invoice_data.columns:
                if col in visible_columns:
                    self.tree.column(col, width=self.tree.column(col, "width"), stretch=True)
                else:
                    self.tree.column(col, width=0, stretch=False)
        
        self.current_page = 1  # Reset to first page
        self.refresh_table()
        
        # Close dialog
        dialog.destroy()
    
    def clear_advanced_filters(self):
        """Clear all advanced filters"""
        for key, var in self.filter_vars.items():
            if hasattr(var, 'set'):
                if key == 'status':
                    var.set("All")
                else:
                    var.set("")
        
        # Clear date entries
        if hasattr(self, 'date_from_entry'):
            self.date_from_entry.delete(0, tk.END)
        if hasattr(self, 'date_to_entry'):
            self.date_to_entry.delete(0, tk.END)
        
        # Reset column visibility
        if hasattr(self, 'column_vars'):
            for col, var in self.column_vars.items():
                var.set(True)
                self.tree.column(col, width=self.tree.column(col, "width"), stretch=True)
        
        # Reset filtered invoices
        self.invoice_data.filtered_invoices = self.invoice_data.invoices.copy()
        self.current_page = 1  # Reset to first page
        self.refresh_table()
    
    def on_search(self, *args):
        """Handle search input changes"""
        query = self.search_var.get()
        # Skip search if placeholder text is present
        if query == self.search_placeholder:
            return
            
        self.invoice_data.search_invoices(query)
        self.current_page = 1  # Reset to first page
        self.refresh_table()
    
    def on_search_button(self):
        """Handle search button click"""
        # Skip search if placeholder text is present
        if self.search_var.get() == self.search_placeholder:
            self.search_entry.focus_set()  # Give focus to entry
            return
            
        self.on_search()
    
    def on_auto_filter(self, event=None):
        """Apply status and date filters automatically when changed"""
        status = self.filter_vars['status'].get()
        
        date_from = None
        if hasattr(self, 'date_from_entry') and self.date_from_entry.get():
            date_from = self.date_from_entry.get_date()
        
        date_to = None
        if hasattr(self, 'date_to_entry') and self.date_to_entry.get():
            date_to = self.date_to_entry.get_date()
        
        filters = {
            'status': status,
            'date_from': date_from,
            'date_to': date_to
        }
        
        self.invoice_data.advanced_filter(filters)
        self.current_page = 1  # Reset to first page
        self.refresh_table()
    
    def on_clear_filters(self):
        """Clear all filters"""
        self.filter_vars['status'].set("All")
        
        if hasattr(self, 'date_from_entry'):
            self.date_from_entry.delete(0, tk.END)
        if hasattr(self, 'date_to_entry'):
            self.date_to_entry.delete(0, tk.END)
        
        # Reset search field to placeholder
        self.search_var.set("")
        self.search_entry.insert(0, self.search_placeholder)
        self.search_entry.config(foreground="#888888")
        
        # Clear all advanced filter variables
        for key, var in self.filter_vars.items():
            if hasattr(var, 'set') and key != 'status':
                var.set("")
        
        # Reset column visibility - Add this code to fix the bug
        if hasattr(self, 'column_vars'):
            for col, var in self.column_vars.items():
                var.set(True)
                # Make sure all columns are visible in the treeview
                self.tree.column(col, width=self.tree.column(col, "width"), stretch=True)
        
        self.invoice_data.filtered_invoices = self.invoice_data.invoices.copy()
        self.current_page = 1  # Reset to first page
        self.refresh_table()

    def sort_by_column(self, column):
        """Sort the table by the specified column"""
        # Get current sort direction
        if hasattr(self, 'sort_column') and self.sort_column == column:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = column
            self.sort_reverse = False
        
        # Sort the data
        self.invoice_data.filtered_invoices.sort(
            key=lambda x: x.get(column, ""),
            reverse=self.sort_reverse
        )
        
        # Refresh the table
        self.refresh_table()
    
    def on_prev_page(self):
        """Go to previous page"""
        if self.current_page > 1:
            self.current_page -= 1
            self.refresh_table()
    
    def on_next_page(self):
        """Go to next page"""
        total_invoices = len(self.invoice_data.filtered_invoices)
        max_pages = (total_invoices + self.page_size - 1) // self.page_size
        
        if self.current_page < max_pages:
            self.current_page += 1
            self.refresh_table()
    
    def on_page_size_change(self, event):
        """Handle page size change"""
        if self.page_size_var.get() == "All":
            self.page_size = len(self.invoice_data.filtered_invoices)
        else:
            self.page_size = int(self.page_size_var.get())
        
        self.current_page = 1  # Reset to first page
        self.refresh_table()
    
    def on_invoice_double_click(self, event):
        """Handle double-click on an invoice"""
        item_id = self.tree.focus()
        if not item_id:
            return
        
        # Get the invoice number from the selected row
        item_values = self.tree.item(item_id, "values")
        invoice_number = item_values[self.invoice_data.columns.index("Nr. doc.")]
        
        # Find the invoice in the data
        invoice = None
        for inv in self.invoice_data.filtered_invoices:
            if inv.get("Nr. doc.") == invoice_number:
                invoice = inv
                break
        
        if invoice:
            self.show_invoice_details(invoice)
        else:
            logger.warning(f"Invoice with number {invoice_number} not found")
    
    def show_invoice_details(self, invoice):
        """Show invoice details in a dialog"""
        details_dialog = tk.Toplevel(self)
        details_dialog.title(f"Invoice Details: {invoice.get('Nr. doc.', 'Unknown')}")
        details_dialog.geometry("800x600")
        details_dialog.configure(bg=DARK_BG)
        details_dialog.transient(self)
        details_dialog.grab_set()
        
        # Configure grid
        details_dialog.grid_columnconfigure(0, weight=1)
        details_dialog.grid_rowconfigure(1, weight=1)
        
        # Header
        header_frame = ttk.Frame(details_dialog)
        header_frame.grid(row=0, column=0, padx=20, pady=20, sticky="ew")
        
        header_label = ttk.Label(header_frame, text=f"Invoice: {invoice.get('Nr. doc.', 'Unknown')}", 
                               font=("Segoe UI", 14, "bold"))
        header_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")
        
        # Simple scrollable frame for details
        main_frame = ttk.Frame(details_dialog)
        main_frame.grid(row=1, column=0, padx=20, pady=0, sticky="nsew")
        
        # Create canvas with scrollbar
        canvas = tk.Canvas(main_frame, bg=DARK_BG, highlightthickness=0)
        canvas.pack(side="left", fill="both", expand=True)
        
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollbar.pack(side="right", fill="y")
        canvas.config(yscrollcommand=scrollbar.set)
        
        # Create a frame for the content
        details_frame = ttk.Frame(canvas, style="TFrame")
        details_frame.columnconfigure(1, weight=1)
        
        # Add invoice details
        row = 0
        for key in self.invoice_data.columns + ["Folder Name"]:
            if key in invoice:
                value = str(invoice.get(key, ""))
                
                # Don't skip any fields - show all values even if empty
                label = ttk.Label(details_frame, text=f"{key}:", font=("Segoe UI", 10, "bold"))
                label.grid(row=row, column=0, padx=(10, 20), pady=5, sticky="w")
                
                value_label = ttk.Label(details_frame, text=value, background=ITEM_BG)
                value_label.grid(row=row, column=1, padx=10, pady=5, sticky="w")
                
                # Create a copy button with local variable capturing
                copy_btn = ttk.Button(
                    details_frame, 
                    text="📋", 
                    width=3,
                    command=lambda v=value: self.copy_to_clipboard(v)
                )
                copy_btn.grid(row=row, column=2, padx=5, pady=5)
                
                row += 1
        
        # Create window for the frame
        canvas.create_window((0, 0), window=details_frame, anchor="nw")
        
        # Update scrollregion after the frame changes size
        details_frame.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))
        
        # Function to handle resizing
        def on_canvas_configure(event):
            # Update the width of the window
            canvas.itemconfig(1, width=event.width)
        canvas.bind("<Configure>", on_canvas_configure)
        
        # Add buttons
        button_frame = ttk.Frame(details_dialog)
        button_frame.grid(row=2, column=0, padx=20, pady=20, sticky="ew")
        
        export_button = ttk.Button(
            button_frame, 
            text="Export to Excel", 
            command=lambda: self.export_single_invoice(invoice)
        )
        export_button.grid(row=0, column=0, padx=10, pady=10)
        
        close_button = ttk.Button(button_frame, text="Close", command=details_dialog.destroy)
        close_button.grid(row=0, column=1, padx=10, pady=10)
        
        logger.info(f"Showing details for invoice {invoice.get('Nr. doc.', 'Unknown')}")
    
    def export_single_invoice(self, invoice):
        """Export a single invoice to Excel"""
        output_path = filedialog.asksaveasfilename(
            title="Save Invoice to Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"Invoice_{invoice.get('Nr. doc.', 'Unknown')}.xlsx"
        )
        
        if not output_path:
            return
        
        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            
            # Add headers
            headers = self.invoice_data.columns
            ws.append(headers)
            
            # Add data
            row = [invoice.get(header, "") for header in headers]
            ws.append(row)
            
            # Adjust column widths
            for col_idx, column in enumerate(ws.columns, 1):
                max_length = 0
                column_name = column[0].value
                
                for i, cell in enumerate(column[:100]):
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = max(max_length, len(str(column_name))) + 3
                ws.column_dimensions[chr(64 + col_idx)].width = min(adjusted_width, 35)
            
            # Save the workbook
            wb.save(output_path)
            messagebox.showinfo("Export Successful", f"Invoice exported to {output_path}")
            logger.info(f"Successfully exported invoice {invoice.get('Nr. doc.', 'Unknown')} to {output_path}")
        except Exception as e:
            error_msg = f"Failed to export invoice: {str(e)}"
            messagebox.showerror("Export Error", error_msg)
            logger.error(error_msg)
            traceback.print_exc()
    
    def on_import(self):
        """Import invoices from a folder or ZIP file"""
        folder_path = filedialog.askopenfilename(
            title="Select a folder or a ZIP containing XML invoices",
            filetypes=[("ZIP files", "*.zip"), ("All files", "*.*")]
        )
        
        if not folder_path:
            return
        
        # Check if it's a directory or a ZIP file
        if not os.path.isdir(folder_path) and not folder_path.lower().endswith('.zip'):
            messagebox.showerror("Invalid Selection", "Please select a folder OR a .zip file.")
            return
        
        # Show progress dialog
        progress_dialog = tk.Toplevel(self)
        progress_dialog.title("Importing Invoices")
        progress_dialog.geometry("400x150")
        progress_dialog.configure(bg=DARK_BG)
        progress_dialog.transient(self)
        progress_dialog.grab_set()
        
        # Configure grid
        progress_dialog.grid_columnconfigure(0, weight=1)
        
        # Add progress bar
        progress_label = ttk.Label(progress_dialog, text="Processing files...")
        progress_label.grid(row=0, column=0, padx=20, pady=(20, 10))
        
        progress_bar = ttk.Progressbar(progress_dialog, mode="determinate", length=300)
        progress_bar.grid(row=1, column=0, padx=20, pady=10, sticky="ew")
        progress_bar["value"] = 0
        
        status_var = tk.StringVar(value="Starting import...")
        status_label = ttk.Label(progress_dialog, textvariable=status_var)
        status_label.grid(row=2, column=0, padx=20, pady=(10, 20))
        
        # Define progress callback
        def update_progress(progress, message):
            progress_bar["value"] = progress
            status_var.set(message)
            progress_dialog.update()
        
        # Process in a separate thread
        def process_thread():
            error_message = None
            try:
                logger.info(f"Starting import from {folder_path}")
                new_count = self.invoice_data.process_folder(folder_path, update_progress)
                # Update UI in the main thread
                self.after(100, lambda: self.finish_import(progress_dialog, new_count))
            except Exception as e:
                error_message = str(e)
                logger.error(f"Import error: {error_message}")
                traceback.print_exc()
                # Use a local variable instead of capturing 'e' in the lambda
                self.after(100, lambda: self.show_import_error(progress_dialog, error_message))
        
        import_thread = Thread(target=process_thread)
        import_thread.daemon = True
        import_thread.start()
    
    def finish_import(self, dialog, new_count):
        """Finish the import process"""
        dialog.destroy()
        
        if new_count > 0:
            messagebox.showinfo("Import Successful", f"Successfully imported {new_count} new invoices.")
            logger.info(f"Successfully imported {new_count} new invoices")
        else:
            messagebox.showinfo("Import Complete", "No new invoices found.")
            logger.info("Import completed with no new invoices found")
        
        # Refresh the table
        self.refresh_table()
    
    def show_import_error(self, dialog, error_message):
        """Show import error"""
        dialog.destroy()
        messagebox.showerror("Import Error", f"Failed to import invoices: {error_message}")
    
    def on_export(self):
        """Export invoices to Excel"""
        if not self.invoice_data.filtered_invoices:
            messagebox.showinfo("No Data", "There are no invoices to export.")
            return
        
        output_path = filedialog.asksaveasfilename(
            title="Save Invoices to Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="Invoices.xlsx"
        )
        
        if not output_path:
            return
        
        # Show progress dialog
        progress_dialog = tk.Toplevel(self)
        progress_dialog.title("Exporting Invoices")
        progress_dialog.geometry("400x150")
        progress_dialog.configure(bg=DARK_BG)
        progress_dialog.transient(self)
        progress_dialog.grab_set()
        
        # Configure grid
        progress_dialog.grid_columnconfigure(0, weight=1)
        
        # Add progress bar
        progress_label = ttk.Label(progress_dialog, text="Exporting to Excel...")
        progress_label.grid(row=0, column=0, padx=20, pady=(20, 10))
        
        progress_bar = ttk.Progressbar(progress_dialog, mode="indeterminate", length=300)
        progress_bar.grid(row=1, column=0, padx=20, pady=10, sticky="ew")
        progress_bar.start()
        
        status_var = tk.StringVar(value="Exporting invoices...")
        status_label = ttk.Label(progress_dialog, textvariable=status_var)
        status_label.grid(row=2, column=0, padx=20, pady=(10, 20))
        
        # Process in a separate thread
        def export_thread():
            success = False
            error_message = None
            try:
                logger.info(f"Starting export to {output_path}")
                success = self.invoice_data.export_to_excel(output_path)
                # Update UI in the main thread
                self.after(100, lambda: self.finish_export(progress_dialog, success, output_path))
            except Exception as e:
                error_message = str(e)
                logger.error(f"Export error: {error_message}")
                traceback.print_exc()
                # Use a local variable instead of capturing 'e' in the lambda
                self.after(100, lambda: self.show_export_error(progress_dialog, error_message))
        
        export_thread = Thread(target=export_thread)
        export_thread.daemon = True
        export_thread.start()
    
    def finish_export(self, dialog, success, output_path):
        """Finish the export process"""
        dialog.destroy()
        
        if success:
            messagebox.showinfo("Export Successful", f"Successfully exported invoices to {output_path}")
            logger.info(f"Successfully exported invoices to {output_path}")
        else:
            messagebox.showerror("Export Error", "Failed to export invoices.")
            logger.error("Export failed")
    
    def show_export_error(self, dialog, error_message):
        """Show export error"""
        dialog.destroy()
        messagebox.showerror("Export Error", f"Failed to export invoices: {error_message}")
        
    def show_context_menu(self, event):
        """Show context menu on right-click"""
        # Select the item under the cursor
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self.context_menu.post(event.x_root, event.y_root)
    
    def copy_selection(self, event=None):
        """Copy selected cell value to clipboard"""
        # Get selected item
        selection = self.tree.selection()
        if not selection:
            return
        
        # Get the focused column
        column = self.tree.identify_column(event.x if event else 0)
        column_index = int(column.replace('#', '')) - 1
        
        # If no specific column is focused or invalid index, use the first column
        if column_index < 0 or column_index >= len(self.invoice_data.columns):
            column_index = 0
        
        # Get value from the selected cell
        item = selection[0]
        values = self.tree.item(item, 'values')
        if not values or column_index >= len(values):
            return
        
        # Copy to clipboard
        value = str(values[column_index])
        self.clipboard_clear()
        self.clipboard_append(value)
        
        self.show_copy_message()
    
    def copy_row(self):
        """Copy entire row to clipboard in tab-separated format"""
        selection = self.tree.selection()
        if not selection:
            return
        
        # Get all values from the selected row
        item = selection[0]
        values = self.tree.item(item, 'values')
        if not values:
            return
        
        # Format as tab-separated values for pasting into spreadsheets
        row_text = '\t'.join(str(v) for v in values)
        
        # Copy to clipboard
        self.clipboard_clear()
        self.clipboard_append(row_text)
        
        self.show_copy_message()
    
    def copy_to_clipboard(self, text):
        """Copy text to clipboard and show feedback"""
        self.clipboard_clear()
        self.clipboard_append(text)
        self.show_copy_message()
    
    def show_copy_message(self):
        """Show a brief message that copying was successful"""
        try:
            if hasattr(self, 'copy_message_label'):
                self.copy_message_label.destroy()
        except:
            pass
            
        self.copy_message_label = ttk.Label(self, text="Copied to clipboard!", 
                                          background=ACCENT_COLOR, foreground="white",
                                          padding=(10, 5))
        self.copy_message_label.place(relx=0.5, rely=0.9, anchor="center")
        
        # Remove the message after 1.5 seconds
        self.after(1500, lambda: self.copy_message_label.destroy() if hasattr(self, 'copy_message_label') else None)

# Main entry point
if __name__ == "__main__":
    # Check for required packages
    try:
        import lxml
        try:
            import tkcalendar
        except ImportError:
            print("Installing tkcalendar package...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", "tkcalendar"])
            import tkcalendar
    except ImportError:
        import subprocess
        import sys
        
        logger.info("Installing required packages...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "lxml", "openpyxl", "tkcalendar"])
    
    # Start the application
    app = InvoiceDashboard()
    app.mainloop()
